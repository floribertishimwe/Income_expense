from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from .models import Category,Expense
from django.contrib import messages
import json
from django.core.paginator import Paginator
from django.http import JsonResponse,HttpResponse
from userpreferences.models import UserPreference
import datetime
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def search_expenses(request):
    if request.method =='POST':
            
        search_str = json.loads(request.body).get('searchText')
        expenses= Expense.objects.filter(amount__istartswith=search_str,owner=request.user) | Expense.objects.filter(
            date__istartswith=search_str,owner=request.user)|  Expense.objects.filter(
            description__istartswith=search_str,owner=request.user) |  Expense.objects.filter(
            description__icontains=search_str,owner=request.user) |Expense.objects.filter(
            category__icontains=search_str,owner=request.user)
        data =expenses.values()
        return JsonResponse(list(data) , safe = False)

# Create your views here.
@login_required(login_url='authentication/login')
def index(request):
    categories =Category.objects.all()
    expenses = Expense.objects.filter(owner=request.user)
    paginator = Paginator(expenses, 5)
    page_number = request.GET.get('page')
    page_obj = Paginator.get_page(paginator, page_number)
    currency = UserPreference.objects.get(user=request.user).currency
    context = {
        
        'expenses':expenses,
        'page_obj':page_obj,
        'currency':currency, 
    }
    
    return render(request, 'expenses/index.html',context)
@login_required(login_url='authentication/login')
def add_expense(request):
    categories =Category.objects.all()
    context = {
        'categories':categories,
        'values': request.POST
    }    
    
    if request.method =='GET':
            return render( request,'expenses/add_expense.html',context)

    if request.method == 'POST':
        amount = request.POST['amount']
        
        if not amount:
            messages.error(request,"Amount is required")
            return render( request,'expenses/add_expense.html',context)

    
        description = request.POST['description']
        date = request.POST['expense_date']
        category= request.POST['category']
        
        if not description:
            messages.error(request,"Description is required")
            return render( request,'expenses/add_expense.html',context)

        Expense.objects.create(owner=request.user,amount=amount,date=date,category=category,description=description)
        messages.success(request,"Expense saved successfully.")
        return redirect ('expenses')
@login_required(login_url='authentication/login')
def edit_expense(request,id):
    expense=Expense.objects.get(pk=id)
    categories =Category.objects.all()
    context = {
        'expense':expense,
        'values':expense,
        'categories':categories,
    }
    if request.method == 'GET':
        return render(request,'expenses/edit-expense.html',context )
    
    if request.method == 'POST':
        amount = request.POST['amount']
        
        if not amount:
            messages.error(request,"Amount is required")
            return render( request,'expenses/edit_expense.html',context)

    
        description = request.POST['description']
        date = request.POST['expense_date']
        category= request.POST['category'] 
        if not description:
            messages.error(request,"Description is required")
            return render( request,'expenses/edit_expense.html',context)  
        expense.owner = request.user
        expense.amount = amount
        expense.date = date
        expense.category = category
        expense.description = description
        expense.save()
        messages.success(request,"Expense updated successfully.")
        return redirect('expenses')
    else:
        messages.info(request,'Handling post form')
        return render(request,'expenses/edit-expense.html',context )
    
@login_required(login_url='authentication/login')   
def delete_expense( request,id):
    expense=Expense.objects.get(pk=id)
    expense.delete()
    messages.success(request,'Exepense successfully deleted.')
    return redirect ('expenses')

  
def expense_category_summary(request):
    todays_date=datetime.date.today()
    six_months_ago=todays_date-datetime.timedelta(days=30*6)
    expenses=Expense.objects.filter(owner = request.user,date__gte=six_months_ago,date__lte=todays_date)
    finalrep={}
    def get_category(expense):
            return expense.category
    category_list=list(set(map(get_category,expenses)))
    def get_category_amount(category):
        amount=0
        filtered_by_category=expenses.filter(category=category)
        for item in filtered_by_category:
            amount +=item.amount
        return amount
    for x in expenses:
        for y in category_list:
            finalrep[y]=get_category_amount(y)
        
    return JsonResponse({'expense_category_data':finalrep},safe=False)
    
    
@login_required(login_url='authentication/login')    
def stats_View(request):
    return render(request,'expenses/stats.html')
    
def export_csv(request):
    response=HttpResponse(
          headers={
            "Content-Type": "text/csv",
            "Content-Disposition": 'attachment;filename= "Expenses"'+ str(datetime.datetime.now())+ '.csv'
        },)
    writer=csv.writer(response)
    writer.writerow(['Amount','Description','Category', 'Date'])
    expenses =Expense.objects.filter(owner=request.user)
    for expense in expenses:
        writer.writerow([expense.amount,expense.description,expense.category,expense.date])
    return response   
    
def export_excel(request):
    response =HttpResponse(
        
         headers={
            "Content-Type": "application/vnd.ms-excel",
            "Content-Disposition": 'attachment;filename= "Expenses"'+ str(datetime.datetime.now())+ '.xlsx'
        },)
    wb = openpyxl.Workbook()
    # wb['sheet'].title="Expenses"
    ws = wb.active
    ws.title="Expenses"
    row_num = 0
    headers = ['N','Amount','Description','Category','Date']
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(bottom=Side(border_style='thick'),
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'))

    for col_num,header  in enumerate(headers, 1):
        cell=ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    data_font = Font()
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(bottom=Side(border_style='thick'),
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'))
    rows =Expense.objects.filter(owner=request.user)
    for row_num,row in enumerate(rows, 2):
        ws.cell(row=row_num, column=1,value=row_num-1)
        ws.cell(row=row_num, column=2, value=row.amount)
        ws.cell(row=row_num, column=3, value=row.description)
        ws.cell(row=row_num, column=4, value=row.category) 
        ws.cell(row=row_num, column=5, value=row.date)
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = data_alignment
            cell.border = data_border
    wb.save(response)
    return response

def export_pdf(request):
    response=HttpResponse(
        header={
             "Content-Type": "application/pdf",
            "Content-Disposition": 'attachment;filename= "Expenses"'+ str(datetime.datetime.now())+ '.Pdf'
            
        },)
   
        
        
        
        
    
    
            
        